from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook


class OpenpyxlRowParser:
    def parse(self, content: bytes) -> list[dict[str, str]]:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        if sheet is None:
            return []
        rows_iter = sheet.iter_rows(values_only=True)
        header = [str(cell).strip() if cell is not None else "" for cell in next(rows_iter)]

        rows: list[dict[str, str]] = []
        for raw_row in rows_iter:
            if all(value is None for value in raw_row):
                continue
            rows.append(
                {
                    header[index]: "" if value is None else str(value)
                    for index, value in enumerate(raw_row)
                    if index < len(header) and header[index]
                }
            )
        return rows
